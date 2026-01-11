"""
Excelテンプレート作成スクリプト
エクセルサンプル.xlsxを完全に複製してテンプレートファイルを作成します
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from copy import copy


def create_template(source_path='エクセルサンプル.xlsx', output_path='templates/template.xlsx'):
    """
    元のExcelファイルを完全に複製してテンプレートを作成

    Args:
        source_path: 元のExcelファイルパス
        output_path: 出力先テンプレートファイルパス
    """
    if not os.path.exists(source_path):
        print(f"エラー: {source_path} が見つかりません")
        print("プロジェクトルートに「エクセルサンプル.xlsx」を配置してください")
        return False

    print(f"テンプレート作成開始: {source_path} -> {output_path}")

    try:
        # 元のファイルを読み込み
        source_wb = load_workbook(source_path)

        # 新しいテンプレートを作成
        template_wb = Workbook()
        template_wb.remove(template_wb.active)  # デフォルトシート削除

        # 各シートをコピー
        for sheet_name in source_wb.sheetnames:
            print(f"  シートをコピー中: {sheet_name}")
            source_ws = source_wb[sheet_name]
            template_ws = template_wb.create_sheet(sheet_name)

            # 1. セル値と書式をコピー
            for row in source_ws:
                for cell in row:
                    new_cell = template_ws[cell.coordinate]

                    # 値をコピー
                    if cell.value is not None:
                        new_cell.value = cell.value

                    # 書式をコピー
                    if cell.has_style:
                        # フォント
                        if cell.font:
                            new_cell.font = copy(cell.font)

                        # 罫線
                        if cell.border:
                            new_cell.border = copy(cell.border)

                        # 塗りつぶし
                        if cell.fill:
                            new_cell.fill = copy(cell.fill)

                        # 数値フォーマット
                        if cell.number_format:
                            new_cell.number_format = copy(cell.number_format)

                        # 配置
                        if cell.alignment:
                            new_cell.alignment = copy(cell.alignment)

                        # 保護
                        if cell.protection:
                            new_cell.protection = copy(cell.protection)

            # 2. マージセルをコピー
            for merged_cell_range in source_ws.merged_cells.ranges:
                template_ws.merge_cells(str(merged_cell_range))

            # 3. 列幅をコピー（すべての列プロパティを完全にコピー）
            for col_letter, col_dim in source_ws.column_dimensions.items():
                new_col_dim = template_ws.column_dimensions[col_letter]
                # 各プロパティを安全にコピー
                if col_dim.width is not None:
                    new_col_dim.width = col_dim.width
                new_col_dim.hidden = col_dim.hidden
                new_col_dim.auto_size = col_dim.auto_size
                new_col_dim.bestFit = col_dim.bestFit
                new_col_dim.collapsed = col_dim.collapsed
                new_col_dim.outline_level = col_dim.outline_level

            # 4. 行高をコピー（すべての行プロパティを完全にコピー）
            for row_num, row_dim in source_ws.row_dimensions.items():
                new_row_dim = template_ws.row_dimensions[row_num]
                # 各プロパティを安全にコピー
                if row_dim.height is not None:
                    new_row_dim.height = row_dim.height
                new_row_dim.hidden = row_dim.hidden
                new_row_dim.collapsed = row_dim.collapsed
                new_row_dim.outline_level = row_dim.outline_level

            # 5. シートのプロパティをコピー
            template_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor
            template_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines

        # 出力ディレクトリが存在しない場合は作成
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # テンプレートを保存
        template_wb.save(output_path)
        print(f"✓ テンプレート作成完了: {output_path}")
        return True

    except Exception as e:
        print(f"エラー: テンプレート作成に失敗しました - {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = create_template()
    if success:
        print("\n✓ テンプレートファイルが正常に作成されました")
        print("  次のステップ: バックエンドを起動してください")
        print("  uvicorn main:app --reload --port 8000")
    else:
        print("\n✗ テンプレート作成に失敗しました")
        print("  「エクセルサンプル.xlsx」を backend/ ディレクトリに配置してから再実行してください")
