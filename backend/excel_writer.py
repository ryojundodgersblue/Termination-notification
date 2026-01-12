"""
Excel書き込みモジュール
抽出したデータをExcelテンプレートに書き込みます
"""

import os
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# セルマッピング定義
# シート「１５ (１)」- 貸借対照表（資産の部）
BALANCE_SHEET_ASSETS_MAP = {
    '現金及び預金': ('１５ (１)', 'AE12'),
    '売掛金': ('１５ (１)', 'T14'),
    '未成工事支出金': ('１５ (１)', 'T16'),
    '原材料': ('１５ (１)', 'T17'),
    '材料貯蔵品': ('１５ (１)', 'T17'),
    '立替金': ('１５ (１)', 'T21'),
    '流動資産合計': ('１５ (１)', 'AE23'),
    '建物': ('１５ (１)', 'T26'),
    '構築物': ('１５ (１)', 'AD26'),
    '建物・構築物': ('１５ (１)', 'T28'),
    '機械装置': ('１５ (１)', 'T28'),
    '車両運搬具': ('１５ (１)', 'AD28'),
    '機械・運搬具': ('１５ (１)', 'T30'),
    '工具器具・備品': ('１５ (１)', 'T32'),
    '有形固定資産合計': ('１５ (１)', 'AE39'),
    'ソフトウェア': ('１５ (１)', 'T46'),
    '無形固定資産合計': ('１５ (１)', 'AE47'),
    '出資金': ('１５（２）', 'V8'),
    '投資その他の資産合計': ('１５（２）', 'V10'),
    '固定資産合計': ('１５（２）', 'V11'),
    '資産合計': ('１５（２）', 'V19'),
}

# シート「１５（２）」- 投資その他の資産・負債の部
BALANCE_SHEET_LIABILITIES_MAP = {
    '工事未払金': ('１５（２）', 'V24'),
    '未払金': ('１５（２）', 'V27'),
    '未払法人税等': ('１５（２）', 'V29'),
    '未払消費税等': ('１５（２）', 'V30'),
    '未成工事受入金': ('１５（２）', 'V31'),
    '預り金': ('１５（２）', 'V32'),
    '流動負債合計': ('１５（２）', 'V36'),
    '長期借入金': ('１５（２）', 'V39'),
    '役員等借入金': ('１５（２）', 'V44'),
    '固定負債合計': ('１５（２）', 'V45'),
    '負債合計': ('１５（２）', 'V46'),
}

# シート「１５（３）」- 純資産の部
BALANCE_SHEET_EQUITY_MAP = {
    '資本金': ('１５（３）', 'V4'),
    '繰越利益剰余金': ('１５（３）', 'V15'),
    '利益剰余金合計': ('１５（３）', 'V16'),
    '株主資本合計': ('１５（３）', 'V19'),
    '純資産合計': ('１５（３）', 'V26'),
    '負債・純資産合計': ('１５（３）', 'V28'),
}

# シート「１６（４）」- 損益計算書（売上〜販売費）
INCOME_STATEMENT_MAP = {
    '完成工事高': ('１６（４）', 'S9'),
    '完成工事原価': ('１６（４）', 'S12'),
    '完成工事総利益金額': ('１６（４）', 'S15'),
    '役員報酬': ('１６（４）', 'S18'),
    '給与手当': ('１６（４）', 'S19'),
    '給与手当等': ('１６（４）', 'S19'),
    '法定福利費': ('１６（４）', 'S21'),
    '事務用品費等': ('１６（４）', 'S24'),
    '通信交通費': ('１６（４）', 'S25'),
    '旅費交通費': ('１６（４）', 'S25'),
    '通信費': ('１６（４）', 'S25'),
    '動力用水光熱費': ('１６（４）', 'S26'),
    'リース料': ('１６（４）', 'S27'),
    '広告宣伝費': ('１６（４）', 'S28'),
    '研修諸会費': ('１６（４）', 'S30'),
    '交際費': ('１６（４）', 'S31'),
    '外注費': ('１６（４）', 'S32'),
    '地代家賃': ('１６（４）', 'S33'),
    '賃借料': ('１６（４）', 'S33'),
    '減価償却費': ('１６（４）', 'S34'),
    '会議費': ('１６（４）', 'S35'),
    '租税公課': ('１６（４）', 'S36'),
    '保険料': ('１６（４）', 'S37'),
    '雑費': ('１６（４）', 'S38'),
    '販管費合計': ('１６（４）', 'AH38'),
    '営業損失金額': ('１６（４）', 'S39'),
    '営業利益金額': ('１６（４）', 'S39'),
}

# シート「１６（５）」- 損益計算書（営業外損益）・完成工事原価報告書
NON_OPERATING_MAP = {
    '受取利息': ('１６（５）', 'S4'),
    '受取配当金': ('１６（５）', 'S4'),
    '受取利息・配当金': ('１６（５）', 'S4'),
    '雑収入': ('１６（５）', 'S5'),
    '営業外収益合計': ('１６（５）', 'AH5'),
    '支払利息': ('１６（５）', 'S7'),
    '営業外費用合計': ('１６（５）', 'AH10'),
    '経常利益金額': ('１６（５）', 'S11'),
    '税引前当期純利益': ('１６（５）', 'S20'),
    '法人税・住民税・事業税': ('１６（５）', 'S21'),
    '当期純利益': ('１６（５）', 'S23'),
    '材料費': ('１６（５）', 'S31'),
    '労務費': ('１６（５）', 'S32'),
    '外注加工費': ('１６（５）', 'S34'),
    '経費': ('１６（５）', 'S35'),
}

# 完成工事原価報告書のマッピングを追加（NON_OPERATING_MAPに含めているが、別途定義も可能）
COST_REPORT_MAP = {
    '材料費': ('１６（５）', 'S31'),
    '労務費': ('１６（５）', 'S32'),
    '外注加工費': ('１６（５）', 'S34'),
    '経費': ('１６（５）', 'S35'),
    '完成工事原価': ('１６（５）', 'S37'),
}

# シート「１７（６）」- 株主資本等変動計算書
EQUITY_CHANGE_MAP = {
    '当期首残高_資本金': ('１７（６）', 'N15'),
    '当期首残高_繰越利益剰余金': ('１７（６）', 'AH15'),
    '当期純利益': ('１７（６）', 'AH18'),
    '当期末残高_資本金': ('１７（６）', 'N27'),
    '当期末残高_繰越利益剰余金': ('１７（６）', 'AH27'),
}


def write_to_excel(data: Dict[str, Any], template_path: str, output_path: str) -> str:
    """
    抽出データをExcelテンプレートに書き込み

    Args:
        data: PDF解析で抽出したデータ
        template_path: テンプレートファイルパス
        output_path: 出力先ファイルパス

    Returns:
        出力ファイルパス

    Raises:
        FileNotFoundError: テンプレートが見つからない場合
        Exception: 書き込み処理でエラーが発生した場合
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")

    print(f"Excel書き込み開始: {template_path} -> {output_path}")

    try:
        # テンプレートを読み込み
        wb = load_workbook(template_path)

        # 各データカテゴリを書き込み
        write_count = 0

        # 1. 貸借対照表 - 資産の部
        if 'balance_sheet_assets' in data:
            write_count += write_data_to_sheet(
                wb,
                data['balance_sheet_assets'],
                BALANCE_SHEET_ASSETS_MAP,
                "資産の部"
            )

        # 2. 貸借対照表 - 負債の部
        if 'balance_sheet_liabilities' in data:
            write_count += write_data_to_sheet(
                wb,
                data['balance_sheet_liabilities'],
                BALANCE_SHEET_LIABILITIES_MAP,
                "負債の部"
            )

        # 3. 貸借対照表 - 純資産の部
        if 'balance_sheet_equity' in data:
            write_count += write_data_to_sheet(
                wb,
                data['balance_sheet_equity'],
                BALANCE_SHEET_EQUITY_MAP,
                "純資産の部"
            )

        # 4. 損益計算書
        if 'income_statement' in data:
            write_count += write_data_to_sheet(
                wb,
                data['income_statement'],
                INCOME_STATEMENT_MAP,
                "損益計算書"
            )

        # 5. 営業外損益
        if 'non_operating' in data:
            write_count += write_data_to_sheet(
                wb,
                data['non_operating'],
                NON_OPERATING_MAP,
                "営業外損益"
            )

        # 6. 株主資本等変動計算書
        if 'equity_change' in data:
            write_count += write_data_to_sheet(
                wb,
                data['equity_change'],
                EQUITY_CHANGE_MAP,
                "株主資本等変動計算書"
            )

        # 7. 完成工事原価報告書
        if 'cost_report' in data:
            write_count += write_data_to_sheet(
                wb,
                data['cost_report'],
                COST_REPORT_MAP,
                "完成工事原価報告書"
            )

        # 出力ディレクトリが存在しない場合は作成
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # ファイル保存
        wb.save(output_path)

        print(f"✓ Excel書き込み完了: {write_count}件のデータを書き込みました")
        return output_path

    except Exception as e:
        print(f"Excel書き込みエラー: {str(e)}")
        raise


def write_data_to_sheet(wb, data: Dict[str, Any], mapping: Dict[str, tuple], category_name: str) -> int:
    """
    データをExcelシートに書き込み

    Args:
        wb: Workbookオブジェクト
        data: 書き込むデータ
        mapping: セルマッピング辞書
        category_name: カテゴリ名（ログ用）

    Returns:
        書き込んだデータ件数
    """
    write_count = 0

    for item, value in data.items():
        if item in mapping:
            sheet_name, cell_address = mapping[item]

            # シートが存在するか確認
            if sheet_name not in wb.sheetnames:
                print(f"  警告: シート「{sheet_name}」が見つかりません - {item}をスキップ")
                continue

            try:
                ws = wb[sheet_name]

                # すべての数値について下3桁を除去（1000で割る）
                actual_value = value
                if isinstance(value, (int, float)):
                    actual_value = int(value // 1000)

                # セルへの書き込み（マージセル対応）
                cell = ws[cell_address]
                target_cell = None

                if isinstance(cell, MergedCell):
                    # マージセルの場合、左上のセルに書き込む
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # 左上セルに書き込み
                            target_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                            target_cell.value = actual_value
                            break
                else:
                    # 通常セルの場合
                    target_cell = ws[cell_address]
                    target_cell.value = actual_value

                # すべての数値にカンマ区切りフォーマットを適用
                if target_cell and isinstance(value, (int, float)):
                    target_cell.number_format = '#,##0'

                write_count += 1

                # ログ出力（数値の場合は変換前後を表示）
                if isinstance(value, (int, float)):
                    print(f"  {category_name}: {item} = {value:,} -> {actual_value:,} (下3桁除去) -> {sheet_name}!{cell_address}")
                else:
                    print(f"  {category_name}: {item} = {value} -> {sheet_name}!{cell_address}")
            except Exception as e:
                print(f"  エラー: {item}の書き込みに失敗 ({sheet_name}!{cell_address}): {str(e)}")
        else:
            print(f"  情報: {item}はマッピングに定義されていません")

    return write_count


def calculate_derived_values(wb) -> None:
    """
    計算が必要な項目を算出

    Args:
        wb: Workbookオブジェクト

    計算項目:
    - 完成工事総利益 = 完成工事高 - 完成工事原価
    - 営業利益 = 完成工事総利益 - 販売費及び一般管理費
    - 経常利益 = 営業利益 + 営業外収益 - 営業外費用
    """
    # 必要に応じて実装
    # 現状ではテンプレートに計算式が含まれている前提
    pass


if __name__ == "__main__":
    # テスト用
    test_data = {
        'balance_sheet_assets': {
            '現金及び預金': 5000000,
            '売掛金': 3000000,
        },
        'balance_sheet_liabilities': {
            '工事未払金': 2000000,
        },
        'balance_sheet_equity': {
            '資本金': 10000000,
        },
        'income_statement': {
            '完成工事高': 50000000,
            '完成工事原価': 40000000,
        },
        'non_operating': {
            '受取利息': 50000,
        },
        'equity_change': {},
    }

    template_path = 'エクセルサンプル.xlsx'
    output_path = 'uploads/test_output.xlsx'

    try:
        write_to_excel(test_data, template_path, output_path)
        print("✓ テスト完了")
    except Exception as e:
        print(f"✗ テストエラー: {str(e)}")
